import csv
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Read CSV for symbols/descriptions only
rows = []
with open("mock_portfolio.csv", "r") as f:
    reader = csv.DictReader(f)
    for row in reader:
        if row["Symbol"].strip():
            rows.append(row)

# Fetch live prices, PE ratios, and company names
symbols = [r["Symbol"] for r in rows]
print(f"Fetching live market data for: {symbols}")
live_data = {}
for sym in symbols:
    try:
        ticker = yf.Ticker(sym)
        info = ticker.info
        price = info.get("regularMarketPrice") or info.get("previousClose")
        pe = info.get("trailingPE") or info.get("forwardPE")
        name = info.get("shortName") or info.get("longName") or sym
        live_data[sym] = {"price": price, "pe": pe, "name": name}
        print(f"  {sym}: Price = ${price}, PE = {pe}")
    except Exception as e:
        live_data[sym] = {"price": None, "pe": None, "name": sym}
        print(f"  {sym}: Failed - {e}")

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Live Portfolio"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
total_font = Font(bold=True, size=12)
total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
currency_fmt = '#,##0.00'
thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

# Headers
headers = ["Symbol", "Description", "Quantity", "Price", "Market Value", "Cost Basis",
           "Gain/Loss $", "Gain/Loss %", "PE Ratio", "Weighted PE"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# Data rows
total_market_value = 0
total_cost_basis = 0
row_data = []
for i, r in enumerate(rows, 2):
    sym = r["Symbol"]
    data = live_data[sym]
    price = data["price"]
    pe = data["pe"]
    name = data["name"]

    if price is None:
        continue

    qty = int(r["Quantity"])
    cb = 500.00 if sym == "NFLX" else 2000.00
    mv = price * qty
    gain_loss = mv - cb
    gain_loss_pct = gain_loss / cb if cb != 0 else 0

    total_market_value += mv
    total_cost_basis += cb
    row_data.append({"sym": sym, "mv": mv})

    ws.cell(row=i, column=1, value=sym).font = Font(bold=True)
    ws.cell(row=i, column=2, value=name)
    ws.cell(row=i, column=3, value=qty).alignment = Alignment(horizontal="center")
    ws.cell(row=i, column=4, value=price).number_format = currency_fmt
    ws.cell(row=i, column=5, value=mv).number_format = currency_fmt
    ws.cell(row=i, column=6, value=cb).number_format = currency_fmt

    gl_cell = ws.cell(row=i, column=7, value=gain_loss)
    gl_cell.number_format = '+#,##0.00;-#,##0.00'
    gl_cell.font = Font(bold=True, color="006100" if gain_loss >= 0 else "9C0006")

    glp_cell = ws.cell(row=i, column=8, value=gain_loss_pct)
    glp_cell.number_format = '+0.00%;-0.00%'
    if gain_loss_pct >= 0:
        glp_cell.font = Font(color="006100")
        glp_cell.fill = green_fill
    else:
        glp_cell.font = Font(color="9C0006")
        glp_cell.fill = red_fill

    if pe is not None:
        ws.cell(row=i, column=9, value=round(pe, 2)).number_format = '0.00'
    else:
        ws.cell(row=i, column=9, value="N/A")

    for col in range(1, 11):
        ws.cell(row=i, column=col).border = thin_border

# Weighted PE
weighted_pe_sum = 0
for i, r in enumerate(rows, 2):
    sym = r["Symbol"]
    data = live_data[sym]
    pe = data["pe"]
    # Find the matching row_data entry for mv
    rd = next((d for d in row_data if d["sym"] == sym), None)
    if rd is None:
        continue
    mv = rd["mv"]
    if pe is not None:
        weight = mv / total_market_value
        weighted_contrib = pe * weight
        weighted_pe_sum += weighted_contrib
        ws.cell(row=i, column=10, value=round(weighted_contrib, 4)).number_format = '0.0000'
    else:
        ws.cell(row=i, column=10, value="N/A")

# Totals row
total_row = len(rows) + 2
total_gain_loss = total_market_value - total_cost_basis
total_gain_loss_pct = total_gain_loss / total_cost_basis if total_cost_basis != 0 else 0

for col in range(1, 11):
    cell = ws.cell(row=total_row, column=col)
    cell.fill = total_fill
    cell.font = total_font

ws.cell(row=total_row, column=1, value="TOTAL")
ws.cell(row=total_row, column=5, value=total_market_value).number_format = currency_fmt
ws.cell(row=total_row, column=6, value=total_cost_basis).number_format = currency_fmt

tgl_cell = ws.cell(row=total_row, column=7, value=total_gain_loss)
tgl_cell.number_format = '+#,##0.00;-#,##0.00'
tgl_cell.font = Font(bold=True, size=12, color="006100" if total_gain_loss >= 0 else "9C0006")

tglp_cell = ws.cell(row=total_row, column=8, value=total_gain_loss_pct)
tglp_cell.number_format = '+0.00%;-0.00%'
if total_gain_loss_pct >= 0:
    tglp_cell.font = Font(bold=True, size=12, color="006100")
    tglp_cell.fill = green_fill
else:
    tglp_cell.font = Font(bold=True, size=12, color="9C0006")
    tglp_cell.fill = red_fill

ws.cell(row=total_row, column=9, value="Portfolio PE:").font = total_font
ws.cell(row=total_row, column=9).alignment = Alignment(horizontal="right")
ws.cell(row=total_row, column=10, value=round(weighted_pe_sum, 2)).number_format = '0.00'

# Column widths
for col, w in zip("ABCDEFGHIJ", [10, 24, 10, 12, 14, 14, 14, 13, 12, 14]):
    ws.column_dimensions[col].width = w

wb.save("live_portfolio.xlsx")
print(f"\nSaved to live_portfolio.xlsx")
print(f"Total Portfolio Value: ${total_market_value:,.2f}")
print(f"Total Cost Basis:     ${total_cost_basis:,.2f}")
print(f"Total Gain/Loss:      ${total_gain_loss:,.2f} ({total_gain_loss_pct:+.2%})")
print(f"Weighted Portfolio PE: {weighted_pe_sum:.2f}")
